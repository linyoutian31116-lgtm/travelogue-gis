# -*- coding: utf-8 -*-
from __future__ import annotations

import json
import unittest
from io import BytesIO
from pathlib import Path

import app
from openpyxl import load_workbook


class WorkflowTests(unittest.TestCase):
    def test_heuristic_extraction_and_decisions(self) -> None:
        source = (
            "九月初三，自杭州府發舟，暮過虎丘山，泊於半塘。"
            "翌日遙望天目山，友人往靈隱寺，余未從。"
        )
        project = app.analyze_text(source, "測試遊記")
        by_name = {row["originalName"]: row for row in project["mentions"]}
        self.assertEqual(by_name["杭州府"]["autoDecision"], "visited")
        self.assertEqual(by_name["虎丘山"]["movementType"], "passed")
        self.assertEqual(by_name["半塘"]["movementType"], "stayed")
        self.assertEqual(by_name["天目山"]["autoDecision"], "not_visited")
        self.assertEqual(by_name["靈隱寺"]["movementType"], "other_person")
        self.assertEqual(project["travelDate"], "")
        self.assertTrue(all(row["gisDecision"] == "review" for row in project["mentions"]))

    def test_user_prompt_is_loaded(self) -> None:
        metadata = app.prompt_metadata()
        prompt = app.load_agent_prompt()
        self.assertTrue(metadata["configured"])
        self.assertEqual(metadata["sha256"], "2b28e13436f7bb2b")
        self.assertIn("是否經過", prompt)
        self.assertIn("GIS收錄", prompt)
        self.assertIn("previousActualPlace", prompt)

    def test_sample_project_shape(self) -> None:
        sample = json.loads((Path(app.DATA_ROOT) / "sample_project.json").read_text(encoding="utf-8"))
        self.assertEqual(sample["schema"], "travelogue-gis-project/v2")
        self.assertGreater(len(sample["mentions"]), 50)
        self.assertTrue(all(app.valid_coordinate(row["latitude"], row["longitude"]) for row in sample["mentions"]))
        self.assertTrue(all(row["locationStatus"] in {"locatable", "regional", "relative", "unlocatable", "unverified"} for row in sample["mentions"]))
        self.assertTrue(all("gisDecision" in row for row in sample["mentions"]))

    def test_three_project_library_shape(self) -> None:
        project_root = Path(app.PROJECTS_ROOT)
        manifest = json.loads((project_root / "index.json").read_text(encoding="utf-8"))
        self.assertEqual({row["id"] for row in manifest}, {"zheyou", "yuexi4", "qianyou1"})
        for entry in manifest:
            project = json.loads((project_root / entry["dataFile"]).read_text(encoding="utf-8"))
            self.assertEqual(project["schema"], "travelogue-gis-project/v3")
            self.assertGreater(len(project["sourceText"]), 5000)
            self.assertEqual(len(project["mentions"]), entry["mentionCount"])
            self.assertTrue(all(row["administrativeLevel"] in {"prefecture", "county", "other"} for row in project["mentions"]))
            self.assertTrue(any(row["visitReviewRequired"] for row in project["mentions"]))
            self.assertTrue(any(row["coordinateReviewRequired"] for row in project["mentions"]))

    def test_shidian_evidence_is_reviewable_and_non_destructive(self) -> None:
        project_root = Path(app.PROJECTS_ROOT)
        qian = json.loads((project_root / "qianyou1.json").read_text(encoding="utf-8"))
        yuexi = json.loads((project_root / "yuexi4.json").read_text(encoding="utf-8"))
        qian_evidence = [row for row in qian["mentions"] if row.get("gazetteerEvidence")]
        yuexi_evidence = [row for row in yuexi["mentions"] if row.get("gazetteerEvidence")]
        self.assertGreaterEqual(sum(len(row["gazetteerEvidence"]) for row in qian_evidence), 28)
        self.assertGreaterEqual(sum(len(row["gazetteerEvidence"]) for row in yuexi_evidence), 10)

        dragon_cave = next(row for row in yuexi["mentions"] if row["normalizedName"] == "龍隱洞")
        evidence = dragon_cave["gazetteerEvidence"][0]
        candidate = evidence["candidateCoordinate"]
        self.assertEqual(candidate["liMeters"], 576)
        self.assertEqual(candidate["status"], "candidate_only")
        self.assertNotEqual(
            (dragon_cave["latitude"], dragon_cave["longitude"]),
            (candidate["latitude"], candidate["longitude"]),
        )
        self.assertIsNone(dragon_cave["coordinateDecision"])

        whitewater = next(row for row in qian["mentions"] if row["normalizedName"] == "白水鋪")
        self.assertEqual(whitewater["gazetteerEvidence"][0]["evidenceType"], "historical_itinerary")

    def test_xlsx_export(self) -> None:
        project = json.loads((Path(app.PROJECTS_ROOT) / "qianyou1.json").read_text(encoding="utf-8"))
        raw = app.build_xlsx(project)
        self.assertTrue(raw.startswith(b"PK"))
        self.assertGreater(len(raw), 5000)
        wb = load_workbook(BytesIO(raw), read_only=True)
        self.assertIn("識典證據", wb.sheetnames)
        self.assertGreater(wb["識典證據"].max_row, 20)


if __name__ == "__main__":
    unittest.main()
