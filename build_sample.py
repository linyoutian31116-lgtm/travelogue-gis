# -*- coding: utf-8 -*-
"""Build the bundled Xu Xiake demonstration project from the reviewed workbook."""

from __future__ import annotations

import argparse
import hashlib
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
DEFAULT_WORKSPACE = ROOT.parent


def clean(value: Any) -> str:
    return str(value or "").strip()


def rows(worksheet) -> list[dict[str, Any]]:
    iterator = worksheet.iter_rows(values_only=True)
    headers = [clean(value) for value in next(iterator)]
    return [dict(zip(headers, row)) for row in iterator if any(value is not None for value in row)]


def workbook_path(explicit: str | None) -> Path:
    if explicit:
        return Path(explicit).resolve()
    candidates = list((DEFAULT_WORKSPACE / "outputs").glob("*.xlsx"))
    if not candidates:
        raise FileNotFoundError("outputs 內找不到 Excel。")
    return max(candidates, key=lambda path: path.stat().st_mtime)


def movement_type(status: str) -> str:
    if "到達" in status:
        return "reached"
    if "停留" in status or "住宿" in status or "泊" in status or "停宿" in status:
        return "stayed"
    if "經過" in status:
        return "passed"
    if "遊覽" in status:
        return "visited"
    if "遙望" in status or "眺望" in status:
        return "viewed"
    if "方向" in status or "支路" in status:
        return "direction"
    if "他人" in status:
        return "other_person"
    if "歷史" in status or "回憶" in status:
        return "historical_memory"
    if "提及" in status or "引用" in status:
        return "referenced"
    return "unknown"


def decision(value: str) -> str:
    if value in {"經過", "到達", "是"}:
        return "visited"
    if value in {"非經過", "未經過", "否"}:
        return "not_visited"
    return "uncertain"


def mapped(value: Any, options: dict[str, str], fallback: str) -> str:
    return options.get(clean(value), fallback)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", nargs="?")
    parser.add_argument("--output", default=str(ROOT / "data" / "sample_project.json"))
    args = parser.parse_args()

    source = workbook_path(args.workbook)
    wb = load_workbook(source, read_only=True, data_only=True)
    audit = rows(wb["地名审核"])
    entities = rows(wb["唯一GIS实体"])
    manual = rows(wb["人工判定记录"]) if "人工判定记录" in wb.sheetnames else []

    entity_by_name = {clean(row.get("規範地名")): row for row in entities}
    manual_by_excel_row = {int(row["Excel行"]): row for row in manual if isinstance(row.get("Excel行"), (int, float))}
    mentions: list[dict[str, Any]] = []
    evidence_sentences: list[str] = []

    for audit_index, row in enumerate(audit, start=2):
        lat, lon = row.get("緯度"), row.get("經度")
        if not isinstance(lat, (int, float)) or not isinstance(lon, (int, float)):
            continue
        name = clean(row.get("規範地名")) or clean(row.get("原文地名"))
        entity = entity_by_name.get(name, {})
        automatic = decision(clean(row.get("是否經過")))
        manual_row = manual_by_excel_row.get(audit_index)
        manual_decision = decision(clean(manual_row.get("人工判定"))) if manual_row else None
        status = clean(row.get("經過狀態"))
        evidence = clean(row.get("證據原句"))
        evidence_sentences.append(evidence)
        confidence = "high" if clean(row.get("GIS收錄判定")) == "保留" else "medium"
        coordinate_url = clean(entity.get("坐標來源連結"))
        mentions.append(
            {
                "id": f"xlsx-row-{audit_index}",
                "sequence": len(mentions) + 1,
                "sourceRow": audit_index,
                "dateLabel": clean(row.get("日期")),
                "originalName": clean(row.get("原文地名")),
                "normalizedName": name,
                "placeType": clean(row.get("地名類型")) or "待核地名",
                "gisDecision": mapped(row.get("GIS收錄判定"), {"保留": "keep", "不保留": "discard", "待核": "review"}, "review"),
                "recordLevel": mapped(
                    row.get("記錄層級"),
                    {"核心地名": "core", "路線輔助地標": "route_landmark", "排除項": "excluded", "待核地名": "review"},
                    "review",
                ),
                "evidence": evidence,
                "context": evidence,
                "autoDecision": automatic,
                "manualDecision": manual_decision,
                "movementType": movement_type(status),
                "locationStatus": mapped(
                    row.get("定位狀態"),
                    {"可定位": "locatable", "區域級定位": "regional", "相對位置": "relative", "無法定位": "unlocatable", "未查證": "unverified"},
                    "unverified",
                ),
                "aliasRelation": clean(row.get("簡稱或異名關係")),
                "prefecture": clean(row.get("府級歸屬")),
                "county": clean(row.get("縣級歸屬")),
                "previousActualPlace": clean(row.get("上一實際行程地點")),
                "nextActualPlace": clean(row.get("下一實際行程地點")),
                "adjacencyType": mapped(row.get("鄰接關係"), {"里程": "mileage", "方向": "direction"}, "unknown"),
                "decisionReason": clean(row.get("判定原因")),
                "latitude": float(lat),
                "longitude": float(lon),
                "coordinateSource": clean(row.get("坐標來源")),
                "coordinateEvidence": clean(row.get("坐標證據")),
                "coordinateSourceUrl": coordinate_url,
                "confidence": confidence,
                "reviewRequired": automatic != "visited" or confidence != "high",
            }
        )

    source_text = "\n".join(dict.fromkeys(sentence for sentence in evidence_sentences if sentence))
    payload = {
        "schema": "travelogue-gis-project/v2",
        "projectTitle": "徐霞客《浙遊日記》人工審核示例",
        "travelDate": "",
        "summary": f"由已審核工作簿整理的示例，共 {len(mentions)} 次有坐標地名提及。可再次調整判定與坐標。",
        "sourceText": source_text,
        "sourceSignature": hashlib.sha256(source.read_bytes()).hexdigest()[:16],
        "analysisMode": "reviewed_workbook",
        "externalData": "",
        "researchSources": [],
        "prompt": {
            "configured": True,
            "filename": "Prompt for 地名抽取及繪圖.docx",
            "sha256": "2b28e13436f7bb2b",
            "characters": 4193,
        },
        "generatedAt": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
        "sourceWorkbook": source.name,
        "mentions": mentions,
    }
    destination = Path(args.output)
    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Built {destination} with {len(mentions)} geocoded mentions")


if __name__ == "__main__":
    main()
