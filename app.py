# -*- coding: utf-8 -*-
"""Local web server for the Travelogue GIS review workflow.

The browser never receives the OpenAI API key.  When OPENAI_API_KEY is not
configured, the server falls back to a deliberately conservative local
extractor so that the human-review and mapping workflow remains usable.
"""

from __future__ import annotations

import argparse
import base64
import csv
import hashlib
import io
import json
import math
import mimetypes
import os
import re
import sys
import urllib.error
import urllib.parse
import urllib.request
import webbrowser
import zipfile
from datetime import datetime, timezone
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any
from xml.etree import ElementTree


ROOT = Path(__file__).resolve().parent
WEB_ROOT = ROOT / "web"
DATA_ROOT = ROOT / "data"
PROJECTS_ROOT = DATA_ROOT / "projects"
PROMPT_DOCX_PATH = DATA_ROOT / "Prompt for 地名抽取及繪圖.docx"
MAX_BODY_BYTES = 12 * 1024 * 1024
MAX_TEXT_CHARS = 240_000
DEFAULT_MODEL = "gpt-5.6-terra"
OPENAI_URL = "https://api.openai.com/v1/responses"
NOMINATIM_URL = "https://nominatim.openstreetmap.org/search"

VISITED = "visited"
NOT_VISITED = "not_visited"
UNCERTAIN = "uncertain"
DECISIONS = {VISITED, NOT_VISITED, UNCERTAIN}
TRUSTED_RESEARCH_DOMAINS = [
    "shidianguji.com",
    "chgis.hudci.org",
    "openstreetmap.org",
    "wikidata.org",
    "whgazetteer.org",
    "baidu.com",
]


PLACE_SUFFIXES = (
    "府|州|縣|县|郡|城|鎮|镇|村|莊|庄|驛|驿|關|关|山|嶺|岭|峰|寺|庵|院|"
    "橋|桥|溪|江|河|湖|潭|洞|門|门|塘|渡|港|島|岛|臺|台|亭|祠|宮|宫|"
    "塔|海|灣|湾|寨|谷|峽|峡|口|塢|坞|洲|泉|坡|岡|冈|堤|廟|庙"
)
PLACE_PATTERN = re.compile(rf"[\u3400-\u9fff]{{1,8}}(?:{PLACE_SUFFIXES})")


ANALYSIS_SCHEMA: dict[str, Any] = {
    "type": "object",
    "properties": {
        "projectTitle": {"type": "string"},
        "travelDate": {"type": "string"},
        "summary": {"type": "string"},
        "mentions": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "sequence": {"type": "integer"},
                    "dateLabel": {"type": "string"},
                    "originalName": {"type": "string"},
                    "normalizedName": {"type": "string"},
                    "placeType": {"type": "string"},
                    "gisDecision": {
                        "type": "string",
                        "enum": ["keep", "discard", "review"],
                    },
                    "recordLevel": {
                        "type": "string",
                        "enum": ["core", "route_landmark", "excluded", "review"],
                    },
                    "evidence": {"type": "string"},
                    "context": {"type": "string"},
                    "autoDecision": {
                        "type": "string",
                        "enum": [VISITED, NOT_VISITED, UNCERTAIN],
                    },
                    "movementType": {
                        "type": "string",
                        "enum": [
                            "reached",
                            "passed",
                            "stayed",
                            "visited",
                            "viewed",
                            "direction",
                            "referenced",
                            "other_person",
                            "historical_memory",
                            "unknown",
                        ],
                    },
                    "locationStatus": {
                        "type": "string",
                        "enum": ["locatable", "regional", "relative", "unlocatable", "unverified"],
                    },
                    "aliasRelation": {"type": "string"},
                    "prefecture": {"type": "string"},
                    "county": {"type": "string"},
                    "previousActualPlace": {"type": "string"},
                    "nextActualPlace": {"type": "string"},
                    "adjacencyType": {
                        "type": "string",
                        "enum": ["mileage", "direction", "unknown"],
                    },
                    "decisionReason": {"type": "string"},
                    "latitude": {"type": ["number", "null"]},
                    "longitude": {"type": ["number", "null"]},
                    "coordinateSource": {"type": "string"},
                    "coordinateSourceUrl": {"type": "string"},
                    "coordinateEvidence": {"type": "string"},
                    "confidence": {
                        "type": "string",
                        "enum": ["high", "medium", "low"],
                    },
                },
                "required": [
                    "sequence",
                    "dateLabel",
                    "originalName",
                    "normalizedName",
                    "placeType",
                    "gisDecision",
                    "recordLevel",
                    "evidence",
                    "context",
                    "autoDecision",
                    "movementType",
                    "locationStatus",
                    "aliasRelation",
                    "prefecture",
                    "county",
                    "previousActualPlace",
                    "nextActualPlace",
                    "adjacencyType",
                    "decisionReason",
                    "latitude",
                    "longitude",
                    "coordinateSource",
                    "coordinateSourceUrl",
                    "coordinateEvidence",
                    "confidence",
                ],
                "additionalProperties": False,
            },
        },
    },
    "required": ["projectTitle", "travelDate", "summary", "mentions"],
    "additionalProperties": False,
}


PROMPT_INTEGRATION_RULES = """

【網頁 Agent 結構化輸出補充規則】
1. 上述 Prompt 是主要判定依據；以下英文鍵只是在網頁中的固定欄位映射。
2. gisDecision：保留=keep、不保留=discard、待核=review。
3. recordLevel：核心地名=core、路線輔助地標=route_landmark、排除項=excluded、待核地名=review。
4. movementType：到達=reached、經過=passed、停宿／停泊=stayed、遊覽=visited、遙望=viewed、方向或支路=direction、他人行程=other_person、歷史回憶=historical_memory、一般提及=referenced、無法判定=unknown。
5. autoDecision 只表示是否加入作者路線：reached/passed/stayed/visited 對應 visited；viewed/direction/other_person/historical_memory/referenced 對應 not_visited；unknown 對應 uncertain。
6. locationStatus：可定位=locatable、區域級定位=regional、相對位置=relative、無法定位=unlocatable、未查證=unverified。
7. adjacencyType：明確里程相鄰=mileage、明確方向相鄰=direction、無法判定=unknown。文本先後出現不能單獨證明地理相鄰。
8. previousActualPlace 與 nextActualPlace 只填最近的作者實際主路線地點；不可填遙望、他人行程、支路或一般提及。
9. 搜尋來源時優先使用 Prompt 指定的識典古籍、CHGIS、OpenStreetMap、Wikidata、WHG Gazetteer 和百度百科。坐標若無法唯一核實必須為 null。
10. 每次提及各輸出一筆，保留原文證據；同一規範地名可重複出現，但 GIS 實體後續由程式去重。
11. 只輸出指定 JSON Schema；Excel 與地圖由網頁在人工提交後產生。
"""


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def json_bytes(value: Any) -> bytes:
    return json.dumps(value, ensure_ascii=False, indent=2).encode("utf-8")


def valid_coordinate(lat: Any, lon: Any) -> bool:
    return (
        isinstance(lat, (int, float))
        and not isinstance(lat, bool)
        and isinstance(lon, (int, float))
        and not isinstance(lon, bool)
        and math.isfinite(float(lat))
        and math.isfinite(float(lon))
        and -90 <= float(lat) <= 90
        and -180 <= float(lon) <= 180
    )


def clean_string(value: Any) -> str:
    return str(value or "").strip()


def extract_docx_text(raw: bytes) -> str:
    try:
        with zipfile.ZipFile(io.BytesIO(raw)) as archive:
            xml = archive.read("word/document.xml")
    except (zipfile.BadZipFile, KeyError) as exc:
        raise ValueError("這不是可讀取的 DOCX 文件。") from exc

    root = ElementTree.fromstring(xml)
    namespace = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    paragraphs: list[str] = []
    for paragraph in root.findall(".//w:p", namespace):
        pieces = [node.text or "" for node in paragraph.findall(".//w:t", namespace)]
        rendered = "".join(pieces).strip()
        if rendered:
            paragraphs.append(rendered)
    return "\n".join(paragraphs)


def load_agent_prompt() -> str:
    """Load the user's authoritative prompt from the bundled DOCX."""
    if not PROMPT_DOCX_PATH.exists():
        raise RuntimeError(f"找不到 Agent Prompt：{PROMPT_DOCX_PATH.name}")
    prompt = extract_docx_text(PROMPT_DOCX_PATH.read_bytes()).strip()
    if len(prompt) < 500:
        raise RuntimeError("Agent Prompt 內容過短或無法讀取。")
    return prompt + PROMPT_INTEGRATION_RULES


def prompt_metadata() -> dict[str, Any]:
    if not PROMPT_DOCX_PATH.exists():
        return {"configured": False, "filename": PROMPT_DOCX_PATH.name}
    raw = PROMPT_DOCX_PATH.read_bytes()
    try:
        characters = len(extract_docx_text(raw))
    except ValueError:
        characters = 0
    return {
        "configured": characters > 0,
        "filename": PROMPT_DOCX_PATH.name,
        "sha256": hashlib.sha256(raw).hexdigest()[:16],
        "characters": characters,
    }


def sentence_chunks(text: str) -> list[str]:
    return [chunk.strip() for chunk in re.findall(r"[^。！？!?；;\n]+[。！？!?；;]?", text) if chunk.strip()]


def infer_local_decision(sentence: str, name: str) -> tuple[str, str, str, str]:
    other_person = bool(re.search(r"(?:友|僧|僕|仆|彼|其|仲|靜聞|静闻).{0,8}(?:往|赴|遊|游|至)", sentence))
    reference = bool(re.search(r"(?:遙望|遥望|望見|望见|聞|听|傳|传|稱|称|問|问|指|云|據|据|昔|舊|旧)", sentence))
    stayed = bool(re.search(r"(?:宿|泊|止|憩|留|居|歇)", sentence))
    passed = bool(re.search(r"(?:過|过|經|经|渡|越|穿|循|沿)", sentence))
    reached = bool(re.search(r"(?:抵|至|到|登|入|出|赴|趨|趋|遊|游|發|发|行)", sentence))

    if other_person:
        return NOT_VISITED, "other_person", "句中行動主體可能不是作者本人，需人工覆核。", "low"
    if reference and not (stayed or passed or reached):
        movement = "viewed" if re.search(r"(?:望|遙望|遥望)", sentence) else "referenced"
        return NOT_VISITED, movement, "地名出現在遙望、傳聞、引述或方向性語境中。", "medium"
    if stayed:
        return VISITED, "stayed", "句中有住宿、泊舟或停留動詞。", "medium"
    if passed:
        return VISITED, "passed", "句中有經過、渡越或沿行動詞。", "medium"
    if reached:
        return VISITED, "reached", "句中有抵達、登臨、進出或出發動詞。", "medium"
    return UNCERTAIN, "unknown", f"只依規則無法確認作者是否實際到達{name}。", "low"


def clean_heuristic_name(name: str) -> str:
    prefixes = (
        r"^(?:(?:翌|次|是|即|本|今|昨|明)?日|(?:清)?晨|早|午|暮|夜|晚|"
        r"友人|同行|僧人|舟子|其人|彼人|余|予|吾|我|"
        r"遙望|遥望|望見|望见|發舟|发舟|登舟|下舟|"
        r"遂|乃|又|復|复|仍|即|自|從|从|由|於|于|往|赴|抵|至|到|"
        r"過|过|經|经|出|入|登|宿|泊|望|遊|游|趨|趋|行)+"
    )
    previous = None
    while previous != name:
        previous = name
        name = re.sub(prefixes, "", name)
    return name.strip()


def heuristic_analyze(text: str, requested_title: str) -> dict[str, Any]:
    mentions: list[dict[str, Any]] = []
    seen_spans: set[tuple[int, int]] = set()
    offset = 0
    for sentence in sentence_chunks(text):
        sentence_start = text.find(sentence, offset)
        if sentence_start < 0:
            sentence_start = offset
        offset = sentence_start + len(sentence)
        clause_offset = 0
        for clause in re.split(r"(?<=[，,、：:])", sentence):
            for match in PLACE_PATTERN.finditer(clause):
                absolute_span = (
                    sentence_start + clause_offset + match.start(),
                    sentence_start + clause_offset + match.end(),
                )
                if absolute_span in seen_spans:
                    continue
                seen_spans.add(absolute_span)
                name = clean_heuristic_name(match.group(0))
                if not name:
                    continue
                decision, movement, reason, confidence = infer_local_decision(clause, name)
                mentions.append(
                    {
                        "sequence": len(mentions) + 1,
                        "dateLabel": "",
                        "originalName": name,
                        "normalizedName": name,
                        "placeType": "待核地名",
                        "gisDecision": "review",
                        "recordLevel": "review",
                        "evidence": sentence,
                        "context": sentence,
                        "autoDecision": decision,
                        "movementType": movement,
                        "locationStatus": "unverified",
                        "aliasRelation": "",
                        "prefecture": "",
                        "county": "",
                        "previousActualPlace": "",
                        "nextActualPlace": "",
                        "adjacencyType": "unknown",
                        "decisionReason": reason,
                        "latitude": None,
                        "longitude": None,
                        "coordinateSource": "",
                        "coordinateSourceUrl": "",
                        "coordinateEvidence": "",
                        "confidence": confidence,
                    }
                )
            clause_offset += len(clause)

    if not mentions:
        raise ValueError("本地初步抽取未找到帶常見地名後綴的詞。可使用「手工新增地點」，或設定 OpenAI API 金鑰啟用 Agent。")
    return {
        "projectTitle": requested_title or "未命名遊記",
        "travelDate": "",
        "summary": f"本地規則初步找到 {len(mentions)} 次地名提及；所有結果都應人工覆核。",
        "mentions": mentions,
    }


def extract_openai_output(response: dict[str, Any]) -> str:
    if response.get("status") == "incomplete":
        reason = (response.get("incomplete_details") or {}).get("reason", "未知原因")
        raise RuntimeError(f"Agent 回應不完整：{reason}")
    for item in response.get("output", []):
        if item.get("type") != "message":
            continue
        for content in item.get("content", []):
            if content.get("type") == "refusal":
                raise RuntimeError(clean_string(content.get("refusal")) or "Agent 拒絕處理這段文本。")
            if content.get("type") == "output_text" and content.get("text"):
                return str(content["text"])
    if response.get("output_text"):
        return str(response["output_text"])
    raise RuntimeError("Agent 沒有返回可解析的結構化結果。")


def extract_web_sources(response: dict[str, Any]) -> list[dict[str, str]]:
    """Collect the complete source list returned by Responses web search."""
    sources: list[dict[str, str]] = []
    seen: set[str] = set()
    for item in response.get("output") or []:
        if item.get("type") != "web_search_call":
            continue
        action = item.get("action") or {}
        for source in action.get("sources") or item.get("sources") or []:
            if not isinstance(source, dict):
                continue
            url = clean_string(source.get("url"))
            if not url or url in seen:
                continue
            seen.add(url)
            sources.append({"title": clean_string(source.get("title")) or url, "url": url})
    return sources


def call_openai(
    text: str,
    requested_title: str,
    travel_date: str = "",
    external_data: str = "",
) -> dict[str, Any]:
    api_key = os.environ.get("OPENAI_API_KEY", "").strip()
    if not api_key:
        raise RuntimeError("OPENAI_API_KEY 尚未設定。")
    model = os.environ.get("OPENAI_MODEL", DEFAULT_MODEL).strip() or DEFAULT_MODEL
    user_content = (
        f"專案名稱：{requested_title or '請依文本擬定'}\n"
        f"旅行日期／年代：{travel_date or '未提供，請勿臆測'}\n\n"
        "可選的地方志、考證或外部資料（不是遊記正文，須與正文證據分開）：\n"
        f"{external_data or '未提供'}\n\n"
        f"原始遊記文本：\n{text}"
    )
    payload = {
        "model": model,
        "input": [
            {"role": "system", "content": load_agent_prompt()},
            {"role": "user", "content": user_content},
        ],
        "reasoning": {"effort": "medium"},
        "text": {
            "verbosity": "low",
            "format": {
                "type": "json_schema",
                "name": "travelogue_place_mentions",
                "description": "逐次遊記地名提及、行程判定與可選坐標。",
                "schema": ANALYSIS_SCHEMA,
                "strict": True,
            },
        },
        "store": False,
        "max_output_tokens": 32000,
    }
    web_search_enabled = os.environ.get("OPENAI_WEB_SEARCH", "1").strip().lower() not in {"0", "false", "no"}
    if web_search_enabled:
        payload["tools"] = [
            {
                "type": "web_search",
                "filters": {"allowed_domains": TRUSTED_RESEARCH_DOMAINS},
            }
        ]
        payload["tool_choice"] = "auto"
        payload["include"] = ["web_search_call.action.sources"]
    request = urllib.request.Request(
        OPENAI_URL,
        data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
            "User-Agent": "TravelogueGIS/0.1",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=300) as response:
            raw = json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", "replace")[:1600]
        try:
            parsed = json.loads(detail)
            detail = clean_string((parsed.get("error") or {}).get("message")) or detail
        except json.JSONDecodeError:
            pass
        raise RuntimeError(f"OpenAI API 錯誤（{exc.code}）：{detail}") from exc
    except urllib.error.URLError as exc:
        raise RuntimeError(f"無法連接 OpenAI API：{exc.reason}") from exc

    output_text = extract_openai_output(raw)
    try:
        result = json.loads(output_text)
        result["_researchSources"] = extract_web_sources(raw)
        return result
    except json.JSONDecodeError as exc:
        raise RuntimeError("Agent 回傳的 JSON 無法解析。") from exc


def normalize_analysis(
    result: dict[str, Any],
    source_text: str,
    mode: str,
    requested_date: str = "",
    external_data: str = "",
) -> dict[str, Any]:
    normalized_mentions: list[dict[str, Any]] = []
    for index, item in enumerate(result.get("mentions") or []):
        original = clean_string(item.get("originalName"))
        name = clean_string(item.get("normalizedName")) or original
        if not name:
            continue
        movement = item.get("movementType")
        movement_options = {
            "reached", "passed", "stayed", "visited", "viewed", "direction",
            "referenced", "other_person", "historical_memory", "unknown",
        }
        movement = movement if movement in movement_options else "unknown"
        decision = item.get("autoDecision") if item.get("autoDecision") in DECISIONS else None
        if decision is None:
            if movement in {"reached", "passed", "stayed", "visited"}:
                decision = VISITED
            elif movement in {"viewed", "direction", "referenced", "other_person", "historical_memory"}:
                decision = NOT_VISITED
            else:
                decision = UNCERTAIN
        gis_decision = item.get("gisDecision") if item.get("gisDecision") in {"keep", "discard", "review"} else "review"
        record_level = item.get("recordLevel") if item.get("recordLevel") in {"core", "route_landmark", "excluded", "review"} else "review"
        location_status = item.get("locationStatus") if item.get("locationStatus") in {"locatable", "regional", "relative", "unlocatable", "unverified"} else "unverified"
        adjacency_type = item.get("adjacencyType") if item.get("adjacencyType") in {"mileage", "direction", "unknown"} else "unknown"
        confidence = item.get("confidence") if item.get("confidence") in {"high", "medium", "low"} else "low"
        lat, lon = item.get("latitude"), item.get("longitude")
        if not valid_coordinate(lat, lon) or location_status != "locatable":
            lat, lon = None, None
        normalized_mentions.append(
            {
                "id": f"mention-{index + 1}",
                "sequence": index + 1,
                "dateLabel": clean_string(item.get("dateLabel")),
                "originalName": original or name,
                "normalizedName": name,
                "placeType": clean_string(item.get("placeType")) or "待核地名",
                "gisDecision": gis_decision,
                "recordLevel": record_level,
                "evidence": clean_string(item.get("evidence")),
                "context": clean_string(item.get("context")),
                "autoDecision": decision,
                "manualDecision": None,
                "movementType": movement,
                "locationStatus": location_status,
                "aliasRelation": clean_string(item.get("aliasRelation")),
                "prefecture": clean_string(item.get("prefecture")),
                "county": clean_string(item.get("county")),
                "previousActualPlace": clean_string(item.get("previousActualPlace")),
                "nextActualPlace": clean_string(item.get("nextActualPlace")),
                "adjacencyType": adjacency_type,
                "decisionReason": clean_string(item.get("decisionReason")),
                "latitude": float(lat) if lat is not None else None,
                "longitude": float(lon) if lon is not None else None,
                "coordinateSource": clean_string(item.get("coordinateSource")),
                "coordinateSourceUrl": clean_string(item.get("coordinateSourceUrl")),
                "coordinateEvidence": clean_string(item.get("coordinateEvidence")),
                "confidence": confidence,
                "reviewRequired": (
                    decision != VISITED
                    or gis_decision == "review"
                    or record_level == "review"
                    or confidence != "high"
                    or (gis_decision == "keep" and lat is None)
                ),
            }
        )

    signature = hashlib.sha256(source_text.encode("utf-8")).hexdigest()[:16]
    return {
        "schema": "travelogue-gis-project/v2",
        "projectTitle": clean_string(result.get("projectTitle")) or "未命名遊記",
        "travelDate": clean_string(result.get("travelDate")) or clean_string(requested_date),
        "summary": clean_string(result.get("summary")),
        "sourceText": source_text,
        "externalData": external_data,
        "sourceSignature": signature,
        "analysisMode": mode,
        "prompt": prompt_metadata(),
        "researchSources": result.get("_researchSources") or [],
        "generatedAt": utc_now(),
        "mentions": normalized_mentions,
    }


def analyze_text(
    text: str,
    requested_title: str,
    travel_date: str = "",
    external_data: str = "",
) -> dict[str, Any]:
    if len(text) > MAX_TEXT_CHARS:
        raise ValueError(f"文本超過 {MAX_TEXT_CHARS:,} 字；請先分篇處理。")
    if len(text.strip()) < 5:
        raise ValueError("請先貼入一段遊記文本。")
    if os.environ.get("OPENAI_API_KEY", "").strip():
        result = call_openai(text, requested_title, travel_date, external_data)
        return normalize_analysis(result, text, "agent", travel_date, external_data)
    result = heuristic_analyze(text, requested_title)
    return normalize_analysis(result, text, "heuristic", travel_date, external_data)


def geocode(query: str, limit: int = 5) -> list[dict[str, Any]]:
    params = urllib.parse.urlencode(
        {"q": query, "format": "jsonv2", "addressdetails": 1, "limit": max(1, min(limit, 8))}
    )
    request = urllib.request.Request(
        f"{NOMINATIM_URL}?{params}",
        headers={"User-Agent": "TravelogueGIS/0.1 (local research tool)"},
    )
    try:
        with urllib.request.urlopen(request, timeout=25) as response:
            rows = json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        raise RuntimeError(f"OpenStreetMap 搜尋失敗（{exc.code}）。") from exc
    except urllib.error.URLError as exc:
        raise RuntimeError(f"無法連接 OpenStreetMap：{exc.reason}") from exc
    results = []
    for row in rows:
        try:
            lat, lon = float(row["lat"]), float(row["lon"])
        except (KeyError, TypeError, ValueError):
            continue
        results.append(
            {
                "displayName": clean_string(row.get("display_name")),
                "type": clean_string(row.get("type")),
                "category": clean_string(row.get("category")),
                "latitude": lat,
                "longitude": lon,
                "source": "OpenStreetMap / Nominatim",
                "sourceUrl": f"https://www.openstreetmap.org/{row.get('osm_type', 'node')}/{row.get('osm_id', '')}",
            }
        )
    return results


def build_xlsx(project: dict[str, Any]) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise RuntimeError("目前的 Python 缺少 openpyxl，無法輸出 Excel。") from exc

    wb = Workbook()
    audit = wb.active
    audit.title = "地名審核"
    headers = [
        "順序",
        "日期",
        "原文地名",
        "規範地名",
        "地名類型",
        "GIS收錄判定",
        "記錄層級",
        "經過狀態",
        "自動判定",
        "人工判定",
        "最終判定",
        "定位狀態",
        "簡稱或異名關係",
        "府級歸屬",
        "縣級歸屬",
        "上一實際行程地點",
        "下一實際行程地點",
        "鄰接關係",
        "判定理由",
        "可信度",
        "緯度",
        "經度",
        "坐標來源",
        "坐標證據",
        "坐標來源連結",
        "證據原句",
        "上下文",
        "識典書名",
        "識典原文",
        "識典結論",
        "識典方向",
        "識典里數",
        "識典候選緯度",
        "識典候選經度",
        "識典連結",
    ]
    audit.append(headers)
    decision_labels = {VISITED: "經過", NOT_VISITED: "未經過", UNCERTAIN: "無法判斷"}
    movement_labels = {
        "reached": "到達", "passed": "經過", "stayed": "停宿／停泊", "visited": "遊覽",
        "viewed": "遙望", "direction": "方向或支路", "other_person": "他人行程",
        "historical_memory": "歷史回憶", "referenced": "一般提及", "unknown": "無法判定",
    }
    gis_labels = {"keep": "保留", "discard": "不保留", "review": "待核"}
    record_labels = {
        "core": "核心地名", "route_landmark": "路線輔助地標",
        "excluded": "排除項", "review": "待核地名",
    }
    location_labels = {
        "locatable": "可定位", "regional": "區域級定位", "relative": "相對位置",
        "unlocatable": "無法定位", "unverified": "未查證",
    }
    adjacency_labels = {"mileage": "里程", "direction": "方向", "unknown": "無法判定"}
    mentions = sorted(project.get("mentions") or [], key=lambda row: row.get("sequence", 0))
    for item in mentions:
        final_decision = item.get("finalDecision") or item.get("manualDecision") or item.get("autoDecision")
        gazetteer = next(iter(item.get("gazetteerEvidence") or []), {})
        candidate = gazetteer.get("candidateCoordinate") or {}
        audit.append(
            [
                item.get("sequence"),
                item.get("dateLabel"),
                item.get("originalName"),
                item.get("normalizedName"),
                item.get("placeType"),
                gis_labels.get(item.get("gisDecision"), item.get("gisDecision")),
                record_labels.get(item.get("recordLevel"), item.get("recordLevel")),
                movement_labels.get(item.get("movementType"), item.get("movementType")),
                decision_labels.get(item.get("autoDecision"), item.get("autoDecision")),
                decision_labels.get(item.get("manualDecision"), item.get("manualDecision") or ""),
                decision_labels.get(final_decision, final_decision),
                location_labels.get(item.get("locationStatus"), item.get("locationStatus")),
                item.get("aliasRelation"),
                item.get("prefecture"),
                item.get("county"),
                item.get("previousActualPlace"),
                item.get("nextActualPlace"),
                adjacency_labels.get(item.get("adjacencyType"), item.get("adjacencyType")),
                item.get("decisionReason"),
                item.get("confidence"),
                item.get("latitude"),
                item.get("longitude"),
                item.get("coordinateSource"),
                item.get("coordinateEvidence"),
                item.get("coordinateSourceUrl"),
                item.get("evidence"),
                item.get("context"),
                gazetteer.get("book"),
                gazetteer.get("quote"),
                gazetteer.get("conclusion"),
                gazetteer.get("direction"),
                gazetteer.get("li"),
                candidate.get("latitude"),
                candidate.get("longitude"),
                gazetteer.get("url"),
            ]
        )

    entities = wb.create_sheet("GIS地名實體")
    entities.append(
        [
            "規範地名", "地名類型", "GIS收錄判定", "記錄層級", "定位狀態",
            "府級歸屬", "縣級歸屬", "緯度", "經度", "坐標來源", "坐標證據",
            "坐標來源連結", "文本提及次數", "實際經停次數", "原文字形集合",
        ]
    )
    groups: dict[str, list[dict[str, Any]]] = {}
    for item in mentions:
        if item.get("gisDecision") == "discard":
            continue
        entity_name = clean_string(item.get("normalizedName")) or clean_string(item.get("originalName"))
        if entity_name:
            groups.setdefault(entity_name, []).append(item)
    for entity_name, rows in groups.items():
        ranked = sorted(
            rows,
            key=lambda row: (
                valid_coordinate(row.get("latitude"), row.get("longitude")),
                row.get("confidence") == "high",
                row.get("gisDecision") == "keep",
            ),
            reverse=True,
        )
        best = ranked[0]
        visited_count = sum(
            1
            for row in rows
            if (row.get("finalDecision") or row.get("manualDecision") or row.get("autoDecision")) == VISITED
        )
        entities.append(
            [
                entity_name,
                best.get("placeType"),
                gis_labels.get(best.get("gisDecision"), best.get("gisDecision")),
                record_labels.get(best.get("recordLevel"), best.get("recordLevel")),
                location_labels.get(best.get("locationStatus"), best.get("locationStatus")),
                best.get("prefecture"),
                best.get("county"),
                best.get("latitude"),
                best.get("longitude"),
                best.get("coordinateSource"),
                best.get("coordinateEvidence"),
                best.get("coordinateSourceUrl"),
                len(rows),
                visited_count,
                "、".join(sorted({clean_string(row.get("originalName")) for row in rows if row.get("originalName")})),
            ]
        )

    route = wb.create_sheet("確認行程")
    route.append(
        [
            "路線順序", "原文順序", "日期", "地名", "經過狀態", "記錄層級",
            "府級歸屬", "縣級歸屬", "緯度", "經度", "坐標來源", "坐標證據", "證據原句",
        ]
    )
    route_index = 0
    for item in mentions:
        decision = item.get("finalDecision") or item.get("manualDecision") or item.get("autoDecision")
        if decision != VISITED or not valid_coordinate(item.get("latitude"), item.get("longitude")):
            continue
        route_index += 1
        route.append(
            [
                route_index,
                item.get("sequence"),
                item.get("dateLabel"),
                item.get("normalizedName"),
                movement_labels.get(item.get("movementType"), item.get("movementType")),
                record_labels.get(item.get("recordLevel"), item.get("recordLevel")),
                item.get("prefecture"),
                item.get("county"),
                item.get("latitude"),
                item.get("longitude"),
                item.get("coordinateSource"),
                item.get("coordinateEvidence"),
                item.get("evidence"),
            ]
        )

    research = wb.create_sheet("研究來源")
    research.append(["名稱", "用途", "連結"])
    for source in project.get("researchSources") or []:
        if isinstance(source, dict):
            research.append([
                source.get("name") or source.get("title"),
                source.get("role") or source.get("description"),
                source.get("url"),
            ])

    shidian = wb.create_sheet("識典證據")
    shidian.append([
        "順序", "地名", "資料類型", "書名", "原文", "匹配關係", "歷史隸屬",
        "方向", "里數", "折合米", "基準點", "候選緯度", "候選經度",
        "結論", "可信度", "理由", "識典連結",
    ])
    for item in mentions:
        for evidence in item.get("gazetteerEvidence") or []:
            candidate = evidence.get("candidateCoordinate") or {}
            shidian.append([
                item.get("sequence"), item.get("normalizedName"), evidence.get("evidenceType"),
                evidence.get("book"), evidence.get("quote"), evidence.get("match"),
                evidence.get("jurisdiction"), evidence.get("direction"), evidence.get("li"),
                evidence.get("meters"), evidence.get("base"), candidate.get("latitude"),
                candidate.get("longitude"), evidence.get("conclusion"), evidence.get("confidence"),
                evidence.get("reason"), evidence.get("url"),
            ])

    info = wb.create_sheet("專案資訊")
    info.append(["項目", "內容"])
    info_rows = [
        ("專案名稱", project.get("projectTitle")),
        ("旅行日期／年代", project.get("travelDate")),
        ("摘要", project.get("summary")),
        ("分析模式", project.get("analysisMode")),
        ("Agent Prompt", (project.get("prompt") or {}).get("filename")),
        ("Prompt 雜湊", (project.get("prompt") or {}).get("sha256")),
        ("產生時間", project.get("generatedAt")),
        ("匯出時間", utc_now()),
        ("文本雜湊", project.get("sourceSignature")),
        ("地名提及數", len(project.get("mentions") or [])),
        ("確認路線點", route_index),
        ("研究來源數", len(project.get("researchSources") or [])),
    ]
    for row in info_rows:
        info.append(row)

    header_fill = PatternFill("solid", fgColor="183B46")
    for worksheet in (audit, entities, route, research, info):
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center")
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column in worksheet.columns:
            letter = column[0].column_letter
            width = min(52, max(10, max(len(str(cell.value or "")) for cell in column) + 2))
            worksheet.column_dimensions[letter].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


class AppHandler(BaseHTTPRequestHandler):
    server_version = "TravelogueGIS/0.3"

    def log_message(self, fmt: str, *args: Any) -> None:
        sys.stdout.write("[%s] %s\n" % (self.log_date_time_string(), fmt % args))

    def send_bytes(self, body: bytes, content_type: str, status: int = 200, headers: dict[str, str] | None = None) -> None:
        self.send_response(status)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store" if self.path.startswith("/api/") else "public, max-age=60")
        self.send_header("X-Content-Type-Options", "nosniff")
        for name, value in (headers or {}).items():
            self.send_header(name, value)
        self.end_headers()
        self.wfile.write(body)

    def send_json(self, value: Any, status: int = 200) -> None:
        self.send_bytes(json_bytes(value), "application/json; charset=utf-8", status)

    def send_error_json(self, status: int, message: str) -> None:
        self.send_json({"error": message}, status)

    def read_json(self) -> dict[str, Any]:
        try:
            length = int(self.headers.get("Content-Length", "0"))
        except ValueError as exc:
            raise ValueError("無效的請求大小。") from exc
        if length <= 0 or length > MAX_BODY_BYTES:
            raise ValueError("請求內容為空或過大。")
        raw = self.rfile.read(length)
        try:
            value = json.loads(raw.decode("utf-8"))
        except (UnicodeDecodeError, json.JSONDecodeError) as exc:
            raise ValueError("請求不是有效的 UTF-8 JSON。") from exc
        if not isinstance(value, dict):
            raise ValueError("請求 JSON 必須是物件。")
        return value

    def do_GET(self) -> None:  # noqa: N802
        parsed = urllib.parse.urlparse(self.path)
        if parsed.path == "/api/status":
            manifest_path = PROJECTS_ROOT / "index.json"
            project_count = 0
            if manifest_path.exists():
                try:
                    project_count = len(json.loads(manifest_path.read_text(encoding="utf-8")))
                except (OSError, json.JSONDecodeError, TypeError):
                    project_count = 0
            self.send_json(
                {
                    "ok": True,
                    "agentConfigured": bool(os.environ.get("OPENAI_API_KEY", "").strip()),
                    "model": os.environ.get("OPENAI_MODEL", DEFAULT_MODEL),
                    "webSearchEnabled": os.environ.get("OPENAI_WEB_SEARCH", "1").strip().lower()
                    not in {"0", "false", "no"},
                    "prompt": prompt_metadata(),
                    "sampleAvailable": (DATA_ROOT / "sample_project.json").exists(),
                    "projectCount": project_count,
                    "version": "0.3.0",
                }
            )
            return
        if parsed.path == "/api/projects":
            manifest_path = PROJECTS_ROOT / "index.json"
            if not manifest_path.exists():
                self.send_error_json(404, "文本資料庫尚未建立。")
                return
            self.send_bytes(manifest_path.read_bytes(), "application/json; charset=utf-8")
            return
        if parsed.path.startswith("/api/projects/"):
            project_id = urllib.parse.unquote(parsed.path.removeprefix("/api/projects/")).strip()
            if not re.fullmatch(r"[a-z0-9_-]+", project_id):
                self.send_error_json(400, "專案識別碼無效。")
                return
            manifest_path = PROJECTS_ROOT / "index.json"
            if not manifest_path.exists():
                self.send_error_json(404, "文本資料庫尚未建立。")
                return
            try:
                manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            except (OSError, json.JSONDecodeError) as exc:
                self.send_error_json(500, f"文本資料庫索引損壞：{exc}")
                return
            entry = next((row for row in manifest if row.get("id") == project_id), None)
            if not entry:
                self.send_error_json(404, "找不到這篇文本。")
                return
            project_path = PROJECTS_ROOT / str(entry.get("dataFile") or f"{project_id}.json")
            if not project_path.is_file():
                self.send_error_json(404, "專案資料檔不存在。")
                return
            self.send_bytes(project_path.read_bytes(), "application/json; charset=utf-8")
            return
        if parsed.path == "/api/sample":
            sample_path = DATA_ROOT / "sample_project.json"
            if not sample_path.exists():
                self.send_error_json(404, "示例資料尚未建立。")
                return
            self.send_bytes(sample_path.read_bytes(), "application/json; charset=utf-8")
            return
        self.serve_static(parsed.path)

    def do_POST(self) -> None:  # noqa: N802
        parsed = urllib.parse.urlparse(self.path)
        try:
            payload = self.read_json()
            if parsed.path == "/api/analyze":
                text = clean_string(payload.get("text"))
                title = clean_string(payload.get("projectTitle"))
                travel_date = clean_string(payload.get("travelDate"))
                external_data = clean_string(payload.get("externalData"))
                self.send_json(analyze_text(text, title, travel_date, external_data))
                return
            if parsed.path == "/api/extract-file":
                filename = clean_string(payload.get("filename"))
                encoded = clean_string(payload.get("data"))
                try:
                    raw = base64.b64decode(encoded, validate=True)
                except (ValueError, TypeError) as exc:
                    raise ValueError("上傳文件內容無法解碼。") from exc
                suffix = Path(filename).suffix.lower()
                if suffix == ".docx":
                    text = extract_docx_text(raw)
                elif suffix in {".txt", ".md", ".text"}:
                    text = raw.decode("utf-8-sig")
                else:
                    raise ValueError("目前支援 TXT、MD 與 DOCX。")
                if len(text) > MAX_TEXT_CHARS:
                    raise ValueError(f"文件超過 {MAX_TEXT_CHARS:,} 字；請先分篇處理。")
                self.send_json({"filename": filename, "text": text, "characters": len(text)})
                return
            if parsed.path == "/api/geocode":
                query = clean_string(payload.get("query"))
                if not query:
                    raise ValueError("請輸入要搜尋的地名。")
                self.send_json({"query": query, "results": geocode(query, int(payload.get("limit", 5)))})
                return
            if parsed.path == "/api/export/xlsx":
                project = payload.get("project")
                if not isinstance(project, dict):
                    raise ValueError("缺少專案資料。")
                body = build_xlsx(project)
                self.send_bytes(
                    body,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": 'attachment; filename="travelogue_gis_review.xlsx"'},
                )
                return
            self.send_error_json(404, "找不到這個 API。")
        except ValueError as exc:
            self.send_error_json(400, str(exc))
        except RuntimeError as exc:
            self.send_error_json(502, str(exc))
        except Exception as exc:  # Keep local UI informative without exposing a traceback.
            self.send_error_json(500, f"伺服器處理失敗：{exc}")

    def serve_static(self, request_path: str) -> None:
        relative = "index.html" if request_path in {"", "/"} else request_path.lstrip("/")
        candidate = (WEB_ROOT / relative).resolve()
        try:
            candidate.relative_to(WEB_ROOT.resolve())
        except ValueError:
            self.send_error_json(403, "禁止存取。")
            return
        if not candidate.is_file():
            candidate = WEB_ROOT / "index.html"
        content_type = mimetypes.guess_type(candidate.name)[0] or "application/octet-stream"
        if content_type.startswith("text/") or content_type in {"application/javascript", "application/json"}:
            content_type += "; charset=utf-8"
        self.send_bytes(candidate.read_bytes(), content_type)


def main() -> None:
    parser = argparse.ArgumentParser(description="啟動遊記行程 GIS 本機網頁")
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--port", type=int, default=8765)
    parser.add_argument("--open", action="store_true", help="啟動後開啟瀏覽器")
    args = parser.parse_args()
    server = ThreadingHTTPServer((args.host, args.port), AppHandler)
    url = f"http://{args.host}:{args.port}/"
    print(f"遊記行程 GIS 已啟動：{url}")
    print("按 Ctrl+C 停止。API 金鑰只從伺服器環境變數 OPENAI_API_KEY 讀取。")
    if args.open:
        webbrowser.open(url)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n已停止。")
    finally:
        server.server_close()


if __name__ == "__main__":
    main()
